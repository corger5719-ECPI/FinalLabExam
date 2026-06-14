from datetime import datetime, date
import csv
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference

# ---------------------------------------------------------
# FUNCTION: createChart
# Arguments:
#   csv_path (str) – The file path to the CSV data file.
#   chart_type (str) – The type of chart to create ("line" or "bar").
# Returns:
#   None – Saves an Excel file (final.xlsx) containing the chart.
# ---------------------------------------------------------
def createChart(csv_path, chart_type):

    print("Choose data source:")
    print("1. Original data")
    print("2. Converted data")
    choice = input("Enter 1 or 2: ")

    data_col = 1 if choice == "1" else 2


    dates = []
    values = []

    with open(csv_path, "r") as file:
        reader = csv.reader(file)
        next(reader)
        for row in reader:
            dates.append(row[0])
            values.append(float(row[data_col]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(["Date", "Value"])

    for d, v in zip(dates, values):
        ws.append([d, v])

    chart = LineChart() if chart_type == "line" else BarChart()

    values_ref = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    labels_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(values_ref, titles_from_data=False)
    chart.set_categories(labels_ref)

    chart.x_axis.title = "Date"
    chart.y_axis.title = "Value"

    today = date.today().strftime("%m/%d/%Y")
    chart.title = "corger5719 " + today

    ws.add_chart(chart, "E5")
    wb.save("final.xlsx")

    print("Chart created successfully in final.xlsx")


# ---------------------------------------------------------
# FUNCTION: generateReport
# Arguments:
#   csv_path (str) – The file path to the CSV data file.
# Returns:
#   None – Calls createChart() to generate the requested report.
# ---------------------------------------------------------
def generateReport(csv_path):

    print("Choose chart type:")
    print("1. Line chart")
    print("2. Bar chart")
    choice = input("Enter 1 or 2: ")

    chart_type = "line" if choice == "1" else "bar"
    createChart(csv_path, chart_type)


# ---------------------------------------------------------
# FUNCTION: convertData
# Arguments:
#   fahrenheit (float) – Temperature value in Fahrenheit.
# Returns:
#   str – Converted Celsius temperature formatted with no decimals.
# ---------------------------------------------------------
def convertData(fahrenheit):
    return f"{(fahrenheit - 32) * 5 / 9:.0f}"


# ---------------------------------------------------------
# FUNCTION: insertData
# Arguments:
#   data (str) – A comma‑separated string of values to append.
#   path (str) – The file path to the CSV file (default: ZooData.csv).
# Returns:
#   bool – True if write succeeds, False if an error occurs.
# ---------------------------------------------------------
def insertData(data, path="ZooData.csv"):
    try:
        with open(path, "a") as file:
            file.write(data + "\n")
        return True
    except:
        print("Error: Data could not be written to the file.")
        return False

# ---------------------------------------------------------
# FUNCTION: viewData
# Arguments:
#   path (str) – The file path to the CSV file.
# Returns:
#   None – Prints the contents of the file to the screen.
# ---------------------------------------------------------
def viewData(path):
    try:
        with open(path, "r") as file:
            print("\nReading file from:", path)
            print(file.read())
    except:
        print("Error: File could not be read.")


# ---------------------------------------------------------
# FUNCTION: getInput
# Arguments:
#   None – User is prompted interactively.
# Returns:
#   None – Saves user‑entered data to ZooData.csv.
# ---------------------------------------------------------
def getInput():
    try:
        entries = int(input("How many entries are being entered? \n"))

        for entry in range(entries):
            date_val = input("Enter the date: \n")
            temperature = float(input("Enter the highest temperature in Fahrenheit: \n"))

            convertedValue = convertData(temperature)
            data = date_val + "," + str(temperature) + "," + str(convertedValue)

            saved = insertData(data, "ZooData.csv")

            if saved:
                print("The following data was saved at", datetime.now(), ":", data)

    except:
        print("Error: Invalid data was entered.")



# ---------------------------------------------------------
# FUNCTION: mainMenu
# Arguments:
#   None – User interacts through menu selections.
# Returns:
#   None – Controls program flow until user exits.
# ---------------------------------------------------------
def mainMenu():
    while True:
        print("\nMAIN MENU")
        print("1. Input Data")
        print("2. View Current Data")
        print("3. Generate Report")
        print("4. Exit")

        selection = input("Enter choice: ")

        if selection == "1":
            getInput()                     # ← OLD OPTION 1
        elif selection == "2":
            viewData("ZooData.csv")        # ← OLD OPTION 2
        elif selection == "3":
            generateReport("ZooData.csv")  # ← NEW REPORTING FEATURE
        elif selection == "4":
            print("Goodbye.")
            break
        else:
            print("Invalid option.")

mainMenu()

