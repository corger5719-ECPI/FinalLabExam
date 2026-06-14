print("corger5719")

import csv
import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference

# FUNCTION 1: Ask user for 5 numbers and total them
def askUser():
    total = 0


    # This loop runs exactly 5 times. Each time, it asks the user for a number,
    # converts it to an integer, and adds it to the running total.
    for x in range(5):
        num = int(input(f"Enter number {x+1}:"))
        total += num

    print("The total of your five numbers is:", total)



# FUNCTION 2: Ask for 5 names + incomes and append to CSV
def askIncome():
  
    # This loop collects 5 entries. Each iteration asks for a name and income,
    # then appends a new line to final.csv so the file grows by 5 rows.
    for x in range(5):
        name = input("Enter person's name: ")
        income = input("Enter annual income: ")

        with open("final.csv","a",newline="") as file:
            writer = csv.writer(file)
            writer.writerow([name, income])



# FUNCTION 3: Create Excel pie chart from CSV data
def excelPie():
    wb = Workbook()
    ws = wb.active

    # Read CSV data into Excel sheet
    with open("final.csv", "r") as file:
        reader = csv.reader(file)
        for row in reader:
            # Convert income to int so Excel chart displays correctly
            row[1] = int(row[1])
            ws.append(row)

 
    # Create a PieChart object
    chart = PieChart()

    # Select the income column (values for the pie slices)
    values = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)

    # Select the names column (labels for the slices)
    labels = Reference(ws, min_col=1, min_row=1, max_row=ws.max_row)

    # Add data and labels to the chart
    chart.add_data(values, titles_from_data=False)
    chart.set_categories(labels)

    # Set chart title to StudentID + today's date
    today = datetime.date.today().strftime("%B %d, %Y")
    chart.title = "corger5719 " + today

    # Add chart to Excel sheet
    ws.add_chart(chart, "E5")

    # Save Excel file
    wb.save("final.xlsx")



# FUNCTION 4: Create vertical bar graph using matplotlib
def verticalBar():
    names = []
    incomes = []

    # Read CSV data
    with open("final.csv", "r") as file:
        reader = csv.reader(file)
        for row in reader:
            names.append(row[0])
            incomes.append(int(row[1]))

    # Create bar graph
    plt.bar(names, incomes)

    # Title with corger5719 + date
    today = datetime.date.today().strftime("%B %d, %Y")
    plt.title("corger5719 " + today)

    plt.xlabel("Names")
    plt.ylabel("Income")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()


askUser()
askIncome()
excelPie()
verticalBar()
